import os
from django.conf import settings
import openpyxl

name = ''
articule = ''
barcode = ''
annotacion = ''
model_list = ''
price = ''
dlinna = ''
shirina = ''
visota = ''
massa = ''
# на вход получаю список [0] - полный путь к файлу [1] - имя с расширением для.


def excel_edit(xl_file_name):

    # Получаем больше информации для заполнения
    more_info_list = more_info(annotacion)
    # открываем книгу.
    book = openpyxl.load_workbook(xl_file_name[0], data_only=False)
    try:
        sheet = book.worksheets[4]  # обращаемся к 5-ому листу в шаблоне OZON
        # цикл отчистки файла.
        for row in sheet.iter_rows(min_row=2, max_row=2, max_col=50):
            count = 0
            for cell in row:
                if cell.value == '':
                    break
                elif cell.value != '':
                    sheet[4][count].value = ''
                count += 1
        # проходим в сторону по 2 стоке всех заполненных колонок. максимально
        # до 50ой колонны.
        for row in sheet.iter_rows(min_row=2, max_row=2, max_col=50):
            # счетчик для обращения к индексу обрабатываемой колонны.
            count = 0
            for cell in row:  # проходим по списку.
                # Если значения во второй строке закончились, обрываем работу.
                if cell.value == '':
                    break
                # далее идет куча условий заполнения документа.
                elif cell.value == 'Артикул*':
                    sheet[4][count].value = articule
                elif cell.value == '№':
                    sheet[4][count].value = '1'
                elif cell.value == 'Название товара':
                    sheet[4][count].value = name
                elif cell.value == 'Цена, руб.*':
                    sheet[4][count].value = price
                elif cell.value == 'НДС, %*':
                    sheet[4][count].value = 'Не облагается'
                elif cell.value == 'Штрихкод (Серийный номер / EAN)':
                    sheet[4][count].value = barcode
                elif cell.value == 'Вес в упаковке, г*':
                    sheet[4][count].value = massa
                elif cell.value == 'Длина упаковки, мм*':
                    sheet[4][count].value = dlinna
                elif cell.value == 'Ширина упаковки, мм*':
                    sheet[4][count].value = shirina
                elif cell.value == 'Высота упаковки, мм*':
                    sheet[4][count].value = visota
                elif cell.value == 'Бренд*':
                    sheet[4][count].value = 'EUROZIP'
                elif (cell.value == 'Название модели '
                      '(для объединения в одну карточку)*'):
                    sheet[4][count].value = articule
                elif (cell.value == 'Страна-изготовитель*'
                      or cell.value == 'Страна-изготовитель'):
                    sheet[4][count].value = 'Италия'
                elif cell.value == 'Партномер':
                    sheet[4][count].value = articule
                elif cell.value == 'Гарантийный срок':
                    sheet[4][count].value = '7 Дней'
                elif cell.value == 'Гарантия':
                    sheet[4][count].value = '7 Дней'
                elif cell.value == 'Комплектация':
                    sheet[4][count].value = name + ' - 1шт'
                elif cell.value == 'Вес товара, г':
                    sheet[4][count].value = str(int(massa) - 35)
                elif cell.value == 'Количество заводских упаковок':
                    sheet[4][count].value = '1'
                elif cell.value == 'Тип*':
                    tipe = xl_file_name[1].replace('_', ' ')
                    sheet[4][count].value = tipe
                elif (cell.value == 'Количество в упаковке, шт'
                      or cell.value == 'Количество в упаковке, шт*'):
                    sheet[4][count].value = '1'
                elif (cell.value == 'Класс опасности товара'
                      or cell.value == 'Класс опасности товара*'):
                    sheet[4][count].value = 'Не опасен'
                elif (cell.value == 'Единиц в одном товаре'
                      or cell.value == 'Единиц в одном товаре*'):
                    sheet[4][count].value = '1'
                elif cell.value == 'Целевая аудитория':
                    sheet[4][count].value = 'Взрослая'
                elif cell.value == 'Срок годности в днях*':
                    sheet[4][count].value = '1800'
                elif cell.value == 'Аннотация':
                    sheet[4][count].value = more_info_list[0][:6000]
                elif (cell.value == 'Список совместимых устройств'
                      or cell.value == 'Предназначено для'
                      or cell.value == 'Совместимые модели минимоек'):
                    sheet[4][count].value = model_list
                elif cell.value == 'Поддерживаемые бренды':
                    if ('DeLonghi' in more_info_list[1]
                            or 'Delonghi' in more_info_list[1]):
                        delonghi = ';'.join(more_info_list[1])
                        delonghi = delonghi.replace('Delonghi', "De'Longhi")
                        delonghi = delonghi.replace('DeLonghi', "De'Longhi")
                        delonghi = delonghi.replace('delonghi', "De'Longhi")
                        sheet[4][count].value = delonghi
                    else:
                        sheet[4][count].value = ';'.join(more_info_list[1])
                elif cell.value == 'Материал':
                    sheet[4][count].value = ';'.join(more_info_list[2])
                elif cell.value == 'Цвет товара':
                    sheet[4][count].value = ';'.join(more_info_list[3])
                elif (cell.value == 'Вид запчасти'
                      or cell.value == 'Вид аксессуара бытовой техники'):
                    sheet[4][count].value = more_info_list[4]
                elif cell.value == 'Ключевые слова':
                    sheet[4][count].value = ';'.join(more_info_list[5])
                elif cell.value == 'Размеры, мм':
                    sheet[4][count].value = 'x'.join(more_info_list[6])
                elif cell.value == 'Количество в упаковке, шт.':
                    sheet[4][count].value = '1'
                elif cell.value == 'Объем, мл':
                    sheet[4][count].value = more_info_list[7]
                elif cell.value == 'ТН ВЭД коды ЕАЭС':
                    sheet[4][count].value = tn_ved_code(xl_file_name[1])
                elif cell.value == 'Мощность, Вт':
                    sheet[4][count].value = more_info_list[8]

                count += 1
        save_file = os.path.join(xl_file_name[0])
        book.save(str(save_file))
        book.close()
        print('Файл успешно заполнен.')
    except TypeError:
        print('Ошибка при создании файла')
        return 'error'


def more_info(annotacion):
    def get_isdigit(info):
        per = ''
        for element in info:
            if element.isdigit():
                per += str(element)
        return per
    gabarites = []
    material_list = ''
    color = ''
    key_val = ''
    brands_list = ''
    objem = ''
    power = ''
    # Получаем бренды, материал, и цвет из аннотации
    for info in annotacion.split('\n'):
        if 'Совместимость с брендом:' in info:
            brands_list = info[25:].split(', ')
        elif 'Материал:' in info:
            material_list = info[10:].split(', ')
        elif 'Цвет:' in info:
            color = info[6:].split(', ')
        elif ('Длина:' in info
              or 'Ширина:' in info
              or 'Высота:' in info
              or 'Диаметр:' in info):
            per = ''
            for el in info:
                if el.isdigit():
                    per = str(per) + str(el)
            if 'см' in info or 'См' in info or 'СМ' in info:
                per = str(int(float(per) * 10))
            elif 'Диаметр' in info:
                gabarites.append(per)
            gabarites.append(per)
        elif 'Объем:' in info:
            objem = get_isdigit(info)
        elif 'Мощность:' in info:
            power = get_isdigit(info)

    # получаем отфильтрованый список моделей для аннотации. Он добавляет
    # только строки с брендами
    model_filter = ''
    for i in brands_list:
        for model in model_list.split('\n'):
            if i in model:
                model_filter += model + ',\n'
    # получаем вид запчасти, зачастую он первый в имени. его и берем.
    vid_zap = name.split()[0]

    try:
        # получаем ключевые слова, зачастую это слова до "Для и чуть дальше,
        key_val = name.split()
        key_val = key_val[:key_val.index('для')+2]
        key_val.remove('-')  # так же удаляем лишние символы.
    except ValueError:
        pass

    finally:
        annotacion = name + '\n\n' + annotacion + model_filter
        return (annotacion, brands_list, material_list, color, vid_zap,
                key_val, gabarites, objem, power)


def on_confirm(entry1, entry2, entry3, text1, text2, entry_price, en_dlinna,
               en_shirina, en_visota, en_massa):
    global name, articule, barcode, annotacion, model_list, price, dlinna
    global shirina, visota, massa
    name = entry1
    articule = entry2
    barcode = entry3
    price = entry_price
    dlinna = en_dlinna
    shirina = en_shirina
    visota = en_visota
    massa = en_massa
    annotacion = text1
    model_list = text2


# Выбираем обрабатываемый файл, а так же берем наименование для сохранения.
def choice_file_xl(file):
    try:
        xl_file_name = os.path.join(
            settings.BASE_DIR, 'media')
        xl_file_name += '/' + str(file)
        file_name = xl_file_name.split('/')
        # получаем наименование с расширением для book.save(file_name)
        file_name = file_name[-1]
        tip = file_name[:-5]
        return xl_file_name, tip
    except TypeError:
        print('Ошибка при выборе файла!')


def tn_ved_code(tip):
    try:
        cartage = ('8516900000 - Части электрические водонагревателей'
                   ' безынерционных или аккумулирующих, '
                   'электронагревателей погружных; электрооборудования'
                   'обогрева пространства и обогрева грунта, '
                   'электротермических аппаратов '
                   'для ухода за волосами (например, сушилки для волос, '
                   'бигуди, щипцы для горячей завивки) и сушилок '
                   'для рук; электроутюгов; прочих бытовых '
                   'электронагревательных приборов; электрических '
                   'нагревательных'
                   ' сопротивлений, кроме указанных в товарной позиции'
                   ' 8545',  # 0
                   '8450900000 - Машины стиральные, бытовые или для прачечных,'
                   ' включая машины, оснащенные отжимным '
                   'устройством: части',  # 1

                   '8422901000 - Части посудомоечных машин',  # 2

                   '7321900000 - Части к кухонным устройствам для '
                   'приготовления и подогрева пищи',  # 3
                   )

        dikt_code = {
            'Запчасть для водонагревателя': cartage[0],
            'Запчасть для обогревателя': cartage[0],
            'Запчасть для стиральной машины': cartage[1],
            'Запчасть для посудомоечной машины': cartage[2],
            'Запчасть для кофемашины': cartage[3],
            'Запчасть для СВЧ': cartage[3],
            'Аксессуар для кофемашины': cartage[3],
            'Аксессуар для кофеварки': cartage[3],
            'Запчасть для кухонного комбайна': cartage[3],
            'Запчасть для кофеварки': cartage[3],
        }

        return dikt_code[tip]
    except KeyError:
        return


def clean_shablon_dir(file_name) -> None:
    file_name = str(file_name).replace(' ', '_')
    xl_file_name = os.path.join(
        settings.BASE_DIR, 'media', 'ozon_shablons')
    files = os.listdir(xl_file_name)
    if file_name in files:

        del_path = f'{xl_file_name}/{file_name}'
        os.remove(del_path)
