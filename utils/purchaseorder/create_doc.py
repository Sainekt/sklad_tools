from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

GREEN_FILL = PatternFill(
    start_color='32cd32', end_color='32cd32', fill_type='solid'
)
YELLOW_FILL = PatternFill(
    start_color='ffff00', end_color='ffff00', fill_type='solid'
)
GRAY_FILL = PatternFill(
    start_color='cccccc', end_color='cccccc', fill_type='solid'
)

BLUE_FILL = PatternFill(
    start_color='33ccff', end_color='33ccff', fill_type='solid'
)

RED_FILL = PatternFill(
    start_color='ff0000', end_color='ff0000', fill_type='solid'
)

BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_report(data, order):
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = 'new'
    work_sheet['A1'] = f'Отчет по заказу №: {order.name}'
    work_sheet.column_dimensions['A'].width = 5
    work_sheet.column_dimensions['B'].width = 110
    work_sheet.column_dimensions['F'].width = 20

    work_sheet['A9'] = '№'
    work_sheet['B9'] = 'Наименование'
    work_sheet['C9'] = 'Код'
    work_sheet['D9'] = 'Заказано'
    work_sheet['E9'] = 'ФАКТ'
    work_sheet['F9'] = 'Комментарий'

    # Цвет

    work_sheet['A9'].fill = GRAY_FILL
    work_sheet['B9'].fill = GRAY_FILL
    work_sheet['C9'].fill = GRAY_FILL
    work_sheet['D9'].fill = GRAY_FILL
    work_sheet['E9'].fill = GRAY_FILL
    work_sheet['F9'].fill = GRAY_FILL

    # обрамление

    work_sheet['A9'].border = BORDER
    work_sheet['B9'].border = BORDER
    work_sheet['C9'].border = BORDER
    work_sheet['D9'].border = BORDER
    work_sheet['E9'].border = BORDER
    work_sheet['F9'].border = BORDER

    count_positions = len(data)
    row = 10  # start row
    index = 0
    ok_count = 0
    more_count = 0
    less_count = 0
    none_count = 0

    def get_style(work_sheet, row, index, color):
        work_sheet[f'A{row+index}'].fill = color
        work_sheet[f'B{row+index}'].fill = color
        work_sheet[f'C{row+index}'].fill = color
        work_sheet[f'D{row+index}'].fill = color
        work_sheet[f'E{row+index}'].fill = color
        work_sheet[f'F{row+index}'].fill = color
        work_sheet[f'A{row+index}'].border = BORDER
        work_sheet[f'B{row+index}'].border = BORDER
        work_sheet[f'C{row+index}'].border = BORDER
        work_sheet[f'D{row+index}'].border = BORDER
        work_sheet[f'E{row+index}'].border = BORDER
        work_sheet[f'F{row+index}'].border = BORDER

    while index <= count_positions - 1:
        products = data[index]
        if products.fact == products.quantity:
            get_style(work_sheet, row, index, GREEN_FILL)
            ok_count += 1
        elif products.fact > products.quantity:
            get_style(work_sheet, row, index, BLUE_FILL)
            more_count += 1
        elif products.fact == 0:
            get_style(work_sheet, row, index, RED_FILL)
            none_count += 1
        else:
            get_style(work_sheet, row, index, YELLOW_FILL)
            less_count += 1
        work_sheet[f'A{row+index}'] = index + 1
        work_sheet[f'B{row+index}'] = products.product.name
        work_sheet[f'C{row+index}'] = products.product.code
        work_sheet[f'D{row+index}'] = products.quantity
        work_sheet[f'E{row+index}'] = products.fact
        work_sheet[f'F{row+index}'] = products.comment
        index += 1

    # Подведение итогов

    work_sheet['B3'] = (f'Сводка: Из {count_positions} позиций.')
    work_sheet['B3'].fill = GRAY_FILL
    work_sheet['B4'] = (f'Сошлось: {ok_count} позиций.')
    work_sheet['B5'] = (f'Больше: {more_count} позиций.')
    work_sheet['B6'] = (f'Меньше: {less_count} позиций.')
    work_sheet['B7'] = (f'Не прислали: {none_count} позиций.')
    work_sheet['B4'].fill = GREEN_FILL
    work_sheet['B5'].fill = BLUE_FILL
    work_sheet['B6'].fill = YELLOW_FILL
    work_sheet['B7'].fill = RED_FILL

    work_sheet['B3'].border = BORDER
    work_sheet['B4'].border = BORDER
    work_sheet['B5'].border = BORDER
    work_sheet['B6'].border = BORDER
    work_sheet['B7'].border = BORDER

    # Сохранение файла

    save_path = f'media/purchaseorder_doc/Отчет_по_заказу_{order.name}.xlsx'
    file_path = f'purchaseorder_doc/Отчет_по_заказу_{order.name}.xlsx'
    work_book.save(save_path)
    work_book.close()
    order.xl_doc = file_path
    order.save()
